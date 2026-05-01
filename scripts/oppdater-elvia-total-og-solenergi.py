#!/usr/bin/env python3
"""
Oppdaterer budsjettarket og andeler.json til:
1. Total strømforbruk = 5 570 863 kWh (faktisk Elvia-tall, var 6 000 000)
2. Oppvarmings-besparelse P2 = 3 633 610 kWh (var 3 875 000), siden bergvarme
   reduserer 75 % × 75 % av redusert totalforbruk
3. Solenergi via overskuddsdeling lagt til i stromBesp P2 (978 180 kWh × 1.20
   kr/kWh, fordelt etter brøk på 430 andeler)

Skriver:
- Ny Excel: ...26.04.26 oppdatert med Elvia-total og solenergi.xlsx
- Ny andeler.json (stromBesp = oppvarming only — simulator legger til solenergi
  dynamisk via brøk)
"""
import json
from pathlib import Path

import openpyxl

SRC = Path(
    "/Users/oyvindborstad/Library/CloudStorage/Dropbox/Orebakken BRL/"
    "Forslag til budsjett Rehab Orebakken 2026 pakke 1 og 2 26.04.26.xlsx"
)
DST_XLSX = SRC.with_name(
    SRC.stem + " oppdatert med Elvia-total og solenergi.xlsx"
)
JSON_OUT = (
    Path(__file__).resolve().parent.parent / "src" / "data" / "andeler.json"
)

TOTAL_FORBRUK_NY = 5_570_863
OPPVARMING_NY = round(0.75 * TOTAL_FORBRUK_NY)
P2_BESPARELSE_NY_KWH = 500_000 + round(0.75 * OPPVARMING_NY)
P2_BESPARELSE_KR_AR = P2_BESPARELSE_NY_KWH * 1.20
P2_BESPARELSE_GAMMEL_KWH = 3_875_000
SCALE_FAKTOR = P2_BESPARELSE_NY_KWH / P2_BESPARELSE_GAMMEL_KWH

SOL_KWH_PER_AR = 978_180
STROMPRIS = 1.20
SOL_KR_PER_AR = SOL_KWH_PER_AR * STROMPRIS


def main() -> None:
    print(f"Skala for oppvarming: {SCALE_FAKTOR:.4f}")
    print(f"Ny besparelse P2: {P2_BESPARELSE_NY_KWH} kWh = {P2_BESPARELSE_KR_AR:,.0f} kr/år")
    print()

    # Last to ganger: data_only for evaluerte verdier, og uten for skriving.
    wb_data = openpyxl.load_workbook(SRC, data_only=True)
    wb = openpyxl.load_workbook(SRC)  # uten data_only = beholder formler

    # 1) Per andel
    ws_data = wb_data["Per andel"]
    ws = wb["Per andel"]
    andeler = []
    rows_data = list(ws_data.iter_rows(min_row=3, values_only=True))
    rows = list(ws.iter_rows(min_row=3))

    for row_data, row in zip(rows_data, rows):
        if not isinstance(row_data[0], (int, float)):
            continue

        andelsnr = int(row_data[0])
        leil_nr = int(row_data[1])
        adresse = row_data[2]
        areal = float(row_data[3])
        brok = float(row_data[4])
        dagens_fu = float(row_data[5])

        p1_ny_fu = float(row_data[6])
        p1_okning = float(row_data[7])
        p1_strom = float(row_data[8])
        p1_skfr = float(row_data[9])
        p1_netto_ar1 = float(row_data[10])
        p1_netto_snitt = float(row_data[11])

        p2_ny_fu = float(row_data[12])
        p2_okning = float(row_data[13])
        p2_strom_old = float(row_data[14])
        p2_skfr = float(row_data[15])
        p2_netto_ar1_old = float(row_data[16])
        p2_netto_snitt_old = float(row_data[17])

        # Skala oppvarming-besparelse til ny total (5.57 mill)
        p2_oppvarming_ny = p2_strom_old * SCALE_FAKTOR  # negativt tall

        # Solenergi (positivt, må trekkes fra netto)
        sol_kr_mnd = (brok * SOL_KR_PER_AR) / 12

        # Excel-versjonen: stromBesp = oppvarming + solar (mer negativ)
        p2_strom_ny_excel = p2_oppvarming_ny - sol_kr_mnd

        # netto_ar1 = brutto_okning - |stromBesp| - |skfr|
        # Endring fra gammel netto: Δ = (|p2_strom_old| - |p2_strom_ny_excel|)
        # = (-p2_strom_old) - (-p2_strom_ny_excel) = p2_strom_ny_excel - p2_strom_old
        delta_strom_excel = p2_strom_ny_excel - p2_strom_old
        # netto endres med -delta_strom (mindre besparelse → høyere netto)
        # men: vi har skalert NED (mindre besparelse) OG lagt til sol (mer besparelse)
        # netto_ny = netto_gammel - delta_strom_excel  (siden netto = brutto - |strom| - skfr)
        # netto = brutto + stromBesp + skfr (signed). Δnetto = Δstrom (signed).
        p2_netto_ar1_ny_excel = p2_netto_ar1_old + delta_strom_excel
        p2_netto_snitt_ny_excel = p2_netto_snitt_old + delta_strom_excel

        # Skriv tilbake til Excel-cellene (overskriver formler med tall)
        row[14].value = p2_strom_ny_excel
        row[16].value = p2_netto_ar1_ny_excel
        row[17].value = p2_netto_snitt_ny_excel

        # JSON-versjonen: stromBesp = oppvarming only (uten sol — simulator legger til)
        # Endring i strøm fra gammel: delta = p2_oppvarming_ny - p2_strom_old (mer positiv = mindre negativ = mindre besparelse)
        delta_strom_json = p2_oppvarming_ny - p2_strom_old
        p2_netto_ar1_json = p2_netto_ar1_old + delta_strom_json
        p2_netto_snitt_json = p2_netto_snitt_old + delta_strom_json

        andeler.append(
            {
                "andelsnr": andelsnr,
                "leilNr": leil_nr,
                "adresse": adresse,
                "areal": areal,
                "brok": brok,
                "dagensFu": dagens_fu,
                "p1": {
                    "nyFu": p1_ny_fu,
                    "okning": p1_okning,
                    "stromBesp": p1_strom,
                    "skfrAr1": p1_skfr,
                    "nettoAr1": p1_netto_ar1,
                    "nettoSnitt": p1_netto_snitt,
                },
                "p2": {
                    "nyFu": p2_ny_fu,
                    "okning": p2_okning,
                    "stromBesp": p2_oppvarming_ny,  # oppvarming only
                    "skfrAr1": p2_skfr,
                    "nettoAr1": p2_netto_ar1_json,
                    "nettoSnitt": p2_netto_snitt_json,
                },
            }
        )

    # Sum-rad i Excel: oppdater til ny sum av strom og netto
    last_row = list(ws.iter_rows())[-1]
    last_row_data = list(ws_data.iter_rows(values_only=True))[-1]
    if last_row_data[0] == "Sum":
        # Beregn nye sums
        new_sum_strom = sum(
            r[14].value
            for r in rows
            if isinstance(r[0].value, (int, float)) and isinstance(r[14].value, (int, float))
        )
        new_sum_netto_ar1 = sum(
            r[16].value
            for r in rows
            if isinstance(r[0].value, (int, float)) and isinstance(r[16].value, (int, float))
        )
        new_sum_netto_snitt = sum(
            r[17].value
            for r in rows
            if isinstance(r[0].value, (int, float)) and isinstance(r[17].value, (int, float))
        )
        last_row[14].value = new_sum_strom
        last_row[16].value = new_sum_netto_ar1
        last_row[17].value = new_sum_netto_snitt

    print(f"Oppdaterte {len(andeler)} andeler i 'Per andel'")

    # 2) Per typologi: kolonne N = 13 (strøm P2), P = 15 (netto), E = 4 (brøk)
    ws_typ_data = wb_data["Per typologi"]
    ws_typ = wb["Per typologi"]
    for row_data, row in zip(
        ws_typ_data.iter_rows(min_row=2, values_only=True),
        ws_typ.iter_rows(min_row=2),
    ):
        if not isinstance(row_data[1], (int, float)):
            continue
        brok = float(row_data[4])
        sol_kr_mnd = (brok * SOL_KR_PER_AR) / 12
        old_strom = float(row_data[13])
        new_oppvarming = old_strom * SCALE_FAKTOR
        new_strom_med_sol = new_oppvarming - sol_kr_mnd
        row[13].value = new_strom_med_sol
        delta_strom = new_strom_med_sol - old_strom
        if isinstance(row_data[15], (int, float)):
            row[15].value = float(row_data[15]) + delta_strom
    print("Oppdaterte 'Per typologi'")

    # 3) Forutsetninger
    ws_for = wb["Forutsetninger"]
    for row in ws_for.iter_rows():
        label = row[0].value
        if not isinstance(label, str):
            continue
        if label == "Totalt strømforbruk (kWh/år)":
            row[1].value = TOTAL_FORBRUK_NY
        elif label == "Besparelse pakke 2 (kWh/år)":
            row[1].value = P2_BESPARELSE_NY_KWH
        elif label == "Besparelse pakke 2 (kr/år)":
            row[1].value = P2_BESPARELSE_KR_AR

    last_row_idx = ws_for.max_row + 2
    ws_for.cell(row=last_row_idx, column=1, value="Solenergi via overskuddsdeling (i tillegg)")
    ws_for.cell(row=last_row_idx + 1, column=1, value="Solproduksjon (kWh/år)")
    ws_for.cell(row=last_row_idx + 1, column=2, value=SOL_KWH_PER_AR)
    ws_for.cell(row=last_row_idx + 2, column=1, value="Verdi (kr/år)")
    ws_for.cell(row=last_row_idx + 2, column=2, value=SOL_KR_PER_AR)
    ws_for.cell(row=last_row_idx + 3, column=1, value="Fordeling")
    ws_for.cell(row=last_row_idx + 3, column=2, value="Etter brøk på 430 andeler via Elhub")
    ws_for.cell(row=last_row_idx + 5, column=1, value="MERK: Total = Elvia-data per 26.02.2026.")
    ws_for.cell(row=last_row_idx + 6, column=1, value="Felles- og privatfordeling presiseres når")
    ws_for.cell(row=last_row_idx + 7, column=1, value="fellesareal-strøm er kartlagt separat.")

    # 4) Skriv Excel
    wb.save(DST_XLSX)
    print(f"\nSkrev Excel:\n  {DST_XLSX}")

    # 5) Skriv ny andeler.json
    JSON_OUT.write_text(
        json.dumps(andeler, ensure_ascii=False, indent=0), encoding="utf-8"
    )
    print(f"Skrev andeler.json med {len(andeler)} andeler")

    # Verifiser snittall
    sum_strom = sum(abs(a["p2"]["stromBesp"]) for a in andeler)
    sum_netto_snitt = sum(a["p2"]["nettoSnitt"] for a in andeler)
    print()
    print(f"P2 strømbesparelse oppvarming snitt (uten sol): {sum_strom/len(andeler):.0f} kr/mnd")
    print(f"P2 netto snitt (uten sol): {sum_netto_snitt/len(andeler):.0f} kr/mnd")
    print(f"P2 netto snitt MED sol: {sum_netto_snitt/len(andeler) - SOL_KR_PER_AR/12/len(andeler):.0f} kr/mnd")


if __name__ == "__main__":
    main()
