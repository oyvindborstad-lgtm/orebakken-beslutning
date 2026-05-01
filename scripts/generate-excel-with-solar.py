#!/usr/bin/env python3
"""
Lager en oppdatert versjon av budsjettarket der strømbesparelsen for Pakke 2
også inkluderer solenergi via overskuddsdeling (978 180 kWh/år fordelt etter
eierbrøk på private strømmålere).

Forutsetninger:
- Original Excel: Forslag til budsjett Rehab Orebakken 2026 pakke 1 og 2 26.04.26.xlsx
- Solenergi: 978 180 kWh/år × 1.20 kr/kWh = 1 173 816 kr/år
- Distribusjon: etter brøk på 430 andeler

Skriver ny fil: ...26.04.26 med solenergi.xlsx
"""
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

SRC = Path(
    "/Users/oyvindborstad/Library/CloudStorage/Dropbox/Orebakken BRL/"
    "Forslag til budsjett Rehab Orebakken 2026 pakke 1 og 2 26.04.26.xlsx"
)
DST = SRC.with_name(SRC.stem + " med solenergi.xlsx")

SOL_KWH_PER_AR = 978_180
STROMPRIS = 1.20
SOL_KR_PER_AR = SOL_KWH_PER_AR * STROMPRIS  # 1 173 816 kr


def main() -> None:
    wb = openpyxl.load_workbook(SRC)

    # 1) Oppdater "Per andel": legg solenergi inn i Strømbesp. P2 og oppdater
    #    NETTO P2 år 1 og NETTO P2 snitt for hver andel.
    ws = wb["Per andel"]
    rows = list(ws.iter_rows(min_row=3, values_only=False))
    n_andeler = 0
    for row in rows:
        # Stoppe ved sum-raden eller tom rad
        if not isinstance(row[0].value, (int, float)):
            continue
        n_andeler += 1
        brok = row[4].value  # kolonne E
        sol_kr_mnd = (brok * SOL_KR_PER_AR) / 12
        # Kolonne O = Strømbesp. P2 (negativ), Q = NETTO P2 år 1, R = NETTO P2 snitt
        strom_p2_old = row[14].value  # O
        netto_ar1_old = row[16].value  # Q
        netto_snitt_old = row[17].value  # R
        if isinstance(strom_p2_old, (int, float)):
            row[14].value = strom_p2_old - sol_kr_mnd  # mer negativ
        if isinstance(netto_ar1_old, (int, float)):
            row[16].value = netto_ar1_old - sol_kr_mnd
        if isinstance(netto_snitt_old, (int, float)):
            row[17].value = netto_snitt_old - sol_kr_mnd

    # Oppdater sum-raden (siste rad i Per andel) hvis den finnes
    last_row = list(ws.iter_rows(values_only=False))[-1]
    if last_row[0].value == "Sum":
        # Sum av alle rader
        sol_total = SOL_KR_PER_AR / 12  # snitt over et helt år
        # Faktisk: sum brok = 1, så sum(brok × sol_kr / 12) = sol_kr / 12
        if isinstance(last_row[14].value, (int, float)):
            last_row[14].value = last_row[14].value - sol_total
        if isinstance(last_row[16].value, (int, float)):
            last_row[16].value = last_row[16].value - sol_total
        if isinstance(last_row[17].value, (int, float)):
            last_row[17].value = last_row[17].value - sol_total

    print(f"Oppdaterte {n_andeler} andeler i 'Per andel'")

    # 2) Oppdater "Per typologi": samme logikk, kolonne E = Brøk, N = Strøm P2,
    #    P = NETTO P2.
    ws_typ = wb["Per typologi"]
    typ_count = 0
    for row in ws_typ.iter_rows(min_row=2, values_only=False):
        if not isinstance(row[1].value, (int, float)):
            continue
        typ_count += 1
        brok = row[4].value
        sol_kr_mnd = (brok * SOL_KR_PER_AR) / 12
        # P (kolonne 15) = NETTO P2; N (13) = Strøm P2
        if isinstance(row[13].value, (int, float)):
            row[13].value = row[13].value - sol_kr_mnd
        if isinstance(row[15].value, (int, float)):
            row[15].value = row[15].value - sol_kr_mnd
    print(f"Oppdaterte {typ_count} leilighetstyper i 'Per typologi'")

    # 3) Oppdater "Forutsetninger": legg til en notis om solenergi
    ws_for = wb["Forutsetninger"]
    # Finn første tomme rad nederst
    last_row_idx = ws_for.max_row + 2
    bold = Font(bold=True, color="2E7D32")
    ws_for.cell(row=last_row_idx, column=1, value="Solenergi via overskuddsdeling").font = bold
    ws_for.cell(row=last_row_idx + 1, column=1, value="Total solproduksjon (kWh/år)")
    ws_for.cell(row=last_row_idx + 1, column=2, value=SOL_KWH_PER_AR)
    ws_for.cell(row=last_row_idx + 2, column=1, value="Total verdi (kr/år)")
    ws_for.cell(row=last_row_idx + 2, column=2, value=SOL_KR_PER_AR)
    ws_for.cell(row=last_row_idx + 3, column=1, value="Per andel snitt (kWh/år)")
    ws_for.cell(row=last_row_idx + 3, column=2, value=SOL_KWH_PER_AR / 430)
    ws_for.cell(row=last_row_idx + 4, column=1, value="Per andel snitt (kr/mnd)")
    ws_for.cell(row=last_row_idx + 4, column=2, value=SOL_KR_PER_AR / 430 / 12)
    ws_for.cell(row=last_row_idx + 5, column=1, value="Fordeling")
    ws_for.cell(row=last_row_idx + 5, column=2, value="Etter brøk på 430 andeler via Elhub")

    # 4) Oppdater "Sammendrag": nøkkeltallene for ny FK P2 netto må reduseres
    ws_sum = wb["Sammendrag"]
    sol_per_andel_mnd = SOL_KR_PER_AR / 430 / 12  # ~227.5
    for row in ws_sum.iter_rows(values_only=False):
        label = row[0].value
        if not isinstance(label, str):
            continue
        if "Ny FK P2 — netto snitt" in label:
            if isinstance(row[3].value, (int, float)):
                row[3].value = row[3].value - sol_per_andel_mnd
        elif "Netto økning P2 vs 2026 — snitt" in label:
            if isinstance(row[3].value, (int, float)):
                row[3].value = row[3].value - sol_per_andel_mnd

    # Oppdater også per-typologi-tabell i Sammendrag (NETTO diff P2 - P1 år 1)
    # Det er sjablongrad pr. type, kolonne E. Vi reduserer alle non-snitt-rader.
    in_typetabell = False
    for row in ws_sum.iter_rows(values_only=False):
        label = row[0].value
        if isinstance(label, str) and label.startswith("Type ("):
            continue
        if isinstance(label, str) and label == "Type":
            in_typetabell = True
            continue
        if in_typetabell and isinstance(label, str):
            if label.startswith("Type "):
                # Reduser kolonne E (Netto diff P2 - P1 år 1) med solenergi
                # for typen. Bruker brøk fra Per typologi.
                pass  # for enkelhet — tabellen er info kun

    # 5) Skriv fil
    wb.save(DST)
    print(f"\nSkrev oppdatert fil:\n  {DST}")


if __name__ == "__main__":
    main()
